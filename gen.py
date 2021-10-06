from openpyxl import Workbook
from openpyxl import load_workbook
import itertools
wb = load_workbook('projets-m2.xlsx',)
print('sheets:',wb.sheetnames)


ws = wb['Form Responses 1']
names=ws['B']
firstnames = ws['C']
sujets = ws['F']
entreprises = ws['D']
pres = ws['R']
stages=ws['M']
phds = ws['N']
#masterlist = (('m1'), ('m2'))
masterlist = (('m2'))
print(masterlist)
def writeRapports(f, n, fn, s, e, p):
    f.write(
"""
 - [[[{0}]]] {1} {2}, _{4}_, link:{{attachmentsdir}}/presentations/{5}.pdf[{5}.pdf] 
""".format(n.value.title().replace(" ", ""), n.value.title(), fn.value.title(), e.value.title().strip(), s.value.capitalize().strip(), str(p.value).strip()))


def writeTableEntry(f, n, fn, s, e, p, stage, phd):
    f.write(
        """
| {0} | {1} | {2} |  _{3}_ | link:{{attachmentsdir}}/presentations/{4}.pdf[{4}.pdf]  | {5} | {6}
""".format(n.value.title(), fn.value.title(), e.value.title().strip(), s.value.strip(), str(p.value).strip(), str(stage.value), str(phd.value)))


for module in ['m2']:
    print('module:',module)
    f = open("modules/"+module+"/partials/presentations.adoc", "w")
    for n, fn, s, e, p in sorted(zip(names, firstnames, sujets, entreprises, pres), key=lambda x: x[0].value):
#            print(str(p.value).strip())
            writeRapports(f, n, fn, s, e, p)
    f.close()
    f = open("modules/"+module+"/partials/projets.adoc", "w")
    f.write('[cols="1,1,2,4,1,1"]\n|===\n')
    f.write('| Nom | Prénom | Entreprise | Sujet | Présentation | Stage | Doctorat \n')
    for n, fn, s, e, p, stage,phd in sorted(zip(names, firstnames, sujets, entreprises, pres,stages,phds), key=lambda x: x[0].value):
            if n.value == 'Nom':
                continue
            print(stage.value, ' ', phd.value )
            writeTableEntry(f, n, fn, s, e, p,stage,phd)
    f.write('\n|===')
    f.close()

encadrants = ws['E']

#emails={'m1':' ','m2':' '}
emails = { 'm2': ' '}
for module in ['m2']:
    for n,e in zip(names,encadrants):
            emails[module]+=e.value.strip()+','

#print('M1:',emails['m1'])
print('M2:',emails['m2'])
